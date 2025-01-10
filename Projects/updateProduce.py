#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet

import openpyxl

loadPath = os.path.join('automate_online-materials', 'produceSales.xlsx')
savePath= os.path.join('Files', 'updatedProduceSales.xlsx')

wb = openpyxl.load_workbook(loadPath)
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07, 
                 'Celery': 1.19,
                 'Lemon': 1.27}

# loop through the rows and update the prices
for rowNum in range(2, sheet.max_row): # skip the first row
    produceName = sheet.cell(row = rowNum, column=1).value # value in column 1 gets stored to produceName
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName] # set value to new price

wb.save(savePath)